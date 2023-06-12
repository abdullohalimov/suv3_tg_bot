from openpyxl import Workbook
from openpyxl.styles import Font
from tempfile import NamedTemporaryFile

def making_excel(data):

    wb = Workbook()

    ws = wb.active

    ws.title = 'So\'rovnoma'

    # cell_range = ws['A1':'I1']

    values = ["ID",	"FIO",	"Region",	"Universitet professori",	"Turk mutaxassisi",	"Bank mutaxassisi",	"Vazirlik mutaxassisi",	"Tashkiliy jarayonlar",	"Fikrlar"]

    ws.append(values)
    ft = Font(bold=True)
    for row in ws["A1:I1"]:
        for cell in row:
            cell.font = ft

    for i in data:
        ws.append(i)

    # with NamedTemporaryFile() as tmp:
    #     wb.save(tmp.name)
    #     tmp.seek(0)
    #     stream = tmp.read()

    wb.save('so\'rovnoma.xlsx')

# making_excel([['ID',	'FIO',	'Region',	'1-savol',	'2-savol',	'3-savol',	'4-savol',	'5-savol',	'Qo\'shimca izoh']])